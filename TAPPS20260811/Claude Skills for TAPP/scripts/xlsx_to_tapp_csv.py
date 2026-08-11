#!/usr/bin/env python3
"""Recover a TAPP CSV from its exported xlsx — the inverse of tapp_to_xlsx.py.

The CSV is the source of truth and the xlsx a generated artifact, so this is a RECOVERY tool, not
part of the normal workflow. It exists because two TAPPs (SEM_TAPP_v10, SEM_Composition_TAPP_v10)
were delivered as xlsx with no CSV counterpart upstream.

Recovery is lossless for content: colour in the xlsx re-encodes the tier already present in
Columns C/D, and the Legends sheet is documentation. Verified by round-tripping
EPMA_TAPP_v13.xlsx against its authored CSV — 97 rows x 29 columns, 0 cell differences.

Output matches the library's conventions exactly: UTF-8 with BOM, CRLF line endings, the TAPP
sheet only. Superscripts and Greek are preserved verbatim.

A recovered file is NOT authoritative. If the real CSV later appears upstream, replace this one
rather than merging into it.

Usage:
    python xlsx_to_tapp_csv.py <file.xlsx> [...]        # writes <file>.csv beside each
    python xlsx_to_tapp_csv.py <file.xlsx> --check      # compare against an existing CSV, write nothing
"""
import csv
import os
import sys

import openpyxl


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "TAPP" if "TAPP" in wb.sheetnames else wb.sheetnames[0]
    rows = [["" if c is None else str(c) for c in r]
            for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    # openpyxl pads short rows to the sheet width; the authored CSVs do the same, so keep as-is
    return rows


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f, lineterminator="\r\n").writerows(rows)


def check(rows, csv_path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        have = list(csv.reader(f))
    diffs = []
    for i in range(max(len(rows), len(have))):
        a = rows[i] if i < len(rows) else []
        b = have[i] if i < len(have) else []
        n = max(len(a), len(b))
        a, b = a + [""] * (n - len(a)), b + [""] * (n - len(b))
        for j, (x, y) in enumerate(zip(a, b)):
            if x.strip() != y.strip():
                diffs.append((i + 1, j + 1, x[:40], y[:40]))
    return diffs


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    do_check = "--check" in sys.argv
    if not args:
        print(__doc__)
        return 1
    for xlsx in args:
        rows = read_xlsx(xlsx)
        out = os.path.splitext(xlsx)[0] + ".csv"
        if do_check:
            if not os.path.exists(out):
                print(f"{os.path.basename(xlsx)}: no CSV to check against")
                continue
            d = check(rows, out)
            print(f"{os.path.basename(xlsx)}: {len(d)} cell difference(s) vs {os.path.basename(out)}")
            for r, c, x, y in d[:5]:
                print(f"    r{r} c{c}: xlsx={x!r} csv={y!r}")
            continue
        write_csv(rows, out)
        print(f"{os.path.basename(xlsx)} -> {os.path.basename(out)}  "
              f"({len(rows)} rows x {len(rows[0])} cols)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
