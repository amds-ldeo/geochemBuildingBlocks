"""Apply ada:componentType string-enum constraints to base BB schemas.

Each ADA file-type building block (image, imageMap, tabularData, collection,
dataCube, document, supDocImage, otherFile) constrains its `ada:componentType`
property to a sealed string enum. The enum lists are derived from the
"Components" worksheet of the canonical spreadsheet
`amds-ldeo/metadata/ADA-AnalyticalMethodsAndAttributes.xlsx`, applying these
mapping rules:

    spreadsheet fileType        ->  base BB
    --------------------------------------------------
    image (isSupplement empty)  ->  image
    image (isSupplement=='supplement')
                                ->  supDocImage
    imageMap                    ->  imageMap
    tabularData, tabularData?   ->  tabularData
    archive                     ->  collection
    dataCube                    ->  dataCube
    document                    ->  document
    document | image            ->  both document AND image
    document | tabularData      ->  both document AND tabularData
    video                       ->  otherFile

Usage:

    # apply enums from cached snapshot (default)
    python tools/apply_componentType_enums.py

    # refresh cache from spreadsheet, then apply
    python tools/apply_componentType_enums.py --refresh \
        --xlsx ../amds-ldeo/metadata/ADA-AnalyticalMethodsAndAttributes.xlsx

The cache `tools/componentType_enum_cache.json` is checked in so the script
can run on any clone without spreadsheet access. Refresh after the
spreadsheet changes; commit the updated cache alongside any schema changes
the apply produces.
"""
import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path
from ruamel.yaml import YAML
from ruamel.yaml.comments import CommentedMap, CommentedSeq

REPO_ROOT = Path(__file__).resolve().parent.parent
GP = REPO_ROOT / "_sources" / "geochemProperties"
DEFAULT_CACHE = REPO_ROOT / "tools" / "componentType_enum_cache.json"
DEFAULT_XLSX = REPO_ROOT.parent.parent / "amds-ldeo" / "metadata" / "ADA-AnalyticalMethodsAndAttributes.xlsx"
DESC = "ADA componentType for this file type, as a single string. Allowed values are derived from the ADA Components mapping (see tools/apply_componentType_enums.py)."

yaml = YAML()
yaml.preserve_quotes = True
yaml.width = 4096
yaml.indent(mapping=2, sequence=4, offset=2)


def derive_from_xlsx(xlsx_path: Path) -> dict:
    """Read the Components sheet and return {bb_name: sorted_componentType_list}."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Components" not in wb.sheetnames:
        raise SystemExit(f"Components sheet not found in {xlsx_path}")
    ws = wb["Components"]
    bb = defaultdict(set)
    for row in ws.iter_rows(min_row=2, values_only=True):
        ct, ft, sup = row
        if ct is None or ft is None:
            continue
        ft = str(ft).strip()
        is_sup = (str(sup).strip().lower() == "supplement") if sup else False
        pfx = "ada:" + str(ct).strip()
        if ft == "image":
            (bb["supDocImage"] if is_sup else bb["image"]).add(pfx)
        elif ft == "imageMap":
            bb["imageMap"].add(pfx)
        elif ft in ("tabularData", "tabularData?"):
            bb["tabularData"].add(pfx)
        elif ft == "archive":
            bb["collection"].add(pfx)
        elif ft == "dataCube":
            bb["dataCube"].add(pfx)
        elif ft == "document":
            bb["document"].add(pfx)
        elif ft == "document | image":
            bb["document"].add(pfx)
            bb["image"].add(pfx)
        elif ft == "document | tabularData":
            bb["document"].add(pfx)
            bb["tabularData"].add(pfx)
        elif ft == "video":
            bb["otherFile"].add(pfx)
        else:
            print(f"  WARN: unhandled fileType={ft!r} for componentType={ct!r}", file=sys.stderr)
    return {k: sorted(v) for k, v in bb.items()}


def replace_componentType_in(node, enum_values, *, only_top_level: bool = False, depth: int = 0) -> int:
    """Walk node; for each dict that has key 'ada:componentType', overwrite the
    value with a string-enum block. If only_top_level is True, only replace at
    depth 1 (immediate properties of the BB root) — used for the collection BB
    so the nested ada:filelist.items.ada:componentType is left as a bare string."""
    n = 0
    if isinstance(node, dict):
        if "ada:componentType" in node and (not only_top_level or depth == 1):
            old = node["ada:componentType"]
            preserved_desc = old.get("description", DESC) if isinstance(old, dict) else DESC
            new = CommentedMap()
            new["type"] = "string"
            enum_seq = CommentedSeq()
            for v in enum_values:
                enum_seq.append(v)
            new["enum"] = enum_seq
            new["description"] = preserved_desc
            node["ada:componentType"] = new
            n += 1
        for v in node.values():
            n += replace_componentType_in(v, enum_values, only_top_level=only_top_level, depth=depth + 1)
    elif isinstance(node, list):
        for v in node:
            n += replace_componentType_in(v, enum_values, only_top_level=only_top_level, depth=depth + 1)
    return n


def apply_to_bb_schemas(mapping: dict) -> int:
    files_changed = 0
    for bb_name, enum_values in sorted(mapping.items()):
        path = GP / bb_name / "schema.yaml"
        if not path.exists():
            print(f"  SKIP {bb_name}: no schema.yaml at {path}")
            continue
        with open(path, encoding="utf-8") as f:
            doc = yaml.load(f)
        only_top = (bb_name == "collection")
        n = replace_componentType_in(doc, enum_values, only_top_level=only_top)
        if n:
            with open(path, "w", encoding="utf-8") as f:
                yaml.dump(doc, f)
            print(f"  {path.relative_to(REPO_ROOT)}: replaced {n} top-level ada:componentType ({len(enum_values)} values)")
            files_changed += 1
        else:
            print(f"  WARN {bb_name}: no top-level ada:componentType in {path.name}")
    return files_changed


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--refresh", action="store_true", help="Re-derive the cache from the spreadsheet before applying.")
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help=f"Path to ADA-AnalyticalMethodsAndAttributes.xlsx (default: {DEFAULT_XLSX})")
    ap.add_argument("--cache", type=Path, default=DEFAULT_CACHE, help=f"Path to JSON cache (default: {DEFAULT_CACHE.relative_to(REPO_ROOT)})")
    args = ap.parse_args()

    if args.refresh:
        if not args.xlsx.exists():
            raise SystemExit(f"xlsx not found: {args.xlsx}")
        mapping = derive_from_xlsx(args.xlsx)
        args.cache.write_text(json.dumps(mapping, indent=2) + "\n", encoding="utf-8")
        print(f"refreshed cache from {args.xlsx} -> {args.cache.relative_to(REPO_ROOT)}")
    else:
        if not args.cache.exists():
            raise SystemExit(f"cache not found ({args.cache}); run with --refresh --xlsx PATH")
        mapping = json.loads(args.cache.read_text(encoding="utf-8"))

    n = apply_to_bb_schemas(mapping)
    print(f"\n{n} BB schemas updated.")


if __name__ == "__main__":
    main()
