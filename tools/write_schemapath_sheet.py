"""Render a workbook's schema-path sidecar CSV as a formatted Excel file for review.

The authoritative, hand-editable mapping lives in `docs/<workbook>.schemapaths.csv` (see
schemapath_io / bootstrap_schemapaths). This tool just presents that CSV as a styled `.xlsx`
(flagged rows highlighted, header frozen) for people who prefer Excel over raw CSV — it does NOT
touch Ruolin's source workbook. Edit the CSV, not the xlsx.

    python tools/write_schemapath_sheet.py docs/EPMA_TAPP_v7.xlsx   # -> docs/EPMA_TAPP_v7.schemapaths.xlsx
"""
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import schemapath_io as sio

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def render(wb_path):
    csv_file = sio.csv_path(wb_path)
    if not os.path.exists(csv_file):
        raise SystemExit(f"no sidecar CSV yet: {os.path.relpath(csv_file, ROOT)} — run bootstrap_schemapaths.py first")
    rows = sio.read(csv_file)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SchemaPaths"
    ws.append(sio.FIELDS)
    for c in ws[1]:
        c.font = Font(bold=True)
    flag_fill = PatternFill("solid", fgColor="FFF3B0")
    for r in rows:
        ws.append([r.get(k, "") for k in sio.FIELDS])
        if (r.get("Source") or "").strip() == "flagged":
            for c in ws[ws.max_row]:
                c.fill = flag_fill
    for col, w in zip("ABCDEFG", (46, 14, 14, 16, 62, 10, 30)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    out = os.path.splitext(wb_path)[0] + ".schemapaths.xlsx"
    wb.save(out)
    n_flag = sum(1 for r in rows if (r.get("Source") or "").strip() == "flagged")
    print(f"{os.path.relpath(out, ROOT)}: {len(rows)} rows, {n_flag} flagged (highlighted)")
    return out


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        raise SystemExit("usage: write_schemapath_sheet.py <workbook.xlsx>")
    render(args[0] if os.path.isabs(args[0]) else os.path.join(ROOT, args[0]))
