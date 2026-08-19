"""Step 2 of the minimal-annotation path: EQUIVALENCE HARNESS.

Proves that (content columns + the overrides sidecar) reproduces exactly what the current
generator derives from the schema-path / implementation-notes annotations — without touching
the generator.

For each tier-column TAPP it runs the REAL build_tapp.route() (which reads the annotations)
to get the ground-truth per-row decisions, then reconstructs each row's decision from content
+ sidecar and asserts they are identical on the three annotation-dependent axes:

  * NAME            of every routed property / parameter
  * ANALYTE COLUMNS emitted for each analyte-template row
  * ANNOTATION SKIP (a row that content-routing WOULD emit but the annotations suppress —
                     inherited base field, protocol description, analyte-identifier row)

Content routing itself (Basic/Advanced tier -> bucket, Data Type -> JSON type, Example ->
enum, mode columns -> analyticalMode) is identical on both sides because it reads the same
content columns, so those need no annotation and are not re-litigated here.

Because build_tapp.build() is a pure function of route()'s result, identical per-row
decisions => identical route() output => identical schema.yaml / *Schema.json. A PASS
therefore means the annotation columns can be dropped in favour of content + sidecar with
zero change to the generated schemas.

    python tools/verify_tapp_overrides.py

Reads overrides from docs/<workbook>.overrides.json (or an `Overrides` worksheet, if present).
"""
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_tapp as b
import _tapp_lib as L
from extract_tapp_overrides import ROOT, TIER_TAPPS, load_rows, load_central_roles, norm_item

ROUTE_P = ("Basic", "Advanced")
ROUTE_A = ("Basic", "Editable", "Advanced")


def load_overrides(docs_xlsx):
    """Prefer a sidecar JSON; fall back to an `Overrides` worksheet in the workbook."""
    side = os.path.join(ROOT, "docs", os.path.splitext(os.path.basename(docs_xlsx))[0] + ".overrides.json")
    if os.path.exists(side):
        with open(side, encoding="utf-8") as f:
            return json.load(f)
    wb = openpyxl.load_workbook(docs_xlsx, data_only=True, read_only=True)
    if "Overrides" not in wb.sheetnames:
        return {}
    ws = wb["Overrides"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr)}
    ov = {}
    for r in rows[1:]:
        item = str(r[col["item"]]).strip() if r and r[col.get("item", 0)] else ""
        if not item:
            continue
        e = {}
        if col.get("name") is not None and r[col["name"]]:
            e["name"] = str(r[col["name"]]).strip()
        if col.get("role") is not None and r[col["role"]]:
            e["role"] = str(r[col["role"]]).strip()
        if col.get("analytecolumn") is not None and r[col["analytecolumn"]]:
            e["analyteColumn"] = [c.strip() for c in str(r[col["analytecolumn"]]).split("|") if c.strip()]
        if e:
            ov[item] = e
    return ov


def ground_truth(tapp):
    """Real route() output keyed by item: routed name, analyte cols, suppressed set, defs."""
    R = b.route()
    name_by_item, analyte_by_item = {}, {}
    routed = set()
    for bucket in ("tapp_prop", "method_param", "method_value", "detail_req", "detail_addl"):
        for rec in R[bucket]:
            name_by_item[rec["item"]] = rec["name"]
            routed.add(rec["item"])
    for rec in R["analyte_cols"]:
        analyte_by_item[rec["item"]] = list(rec.get("cols") or [])
    return name_by_item, analyte_by_item, routed, R["analyte_defs"]


def main():
    central = load_central_roles()
    total = fails = 0
    for tapp in TIER_TAPPS:
        rows = load_rows(tapp)                       # calls b.configure(tapp)
        name_by_item, analyte_by_item, routed, real_defs = ground_truth(tapp)
        ov = load_overrides(b.XLSX)
        base_items = b.CFG["base_items"]
        mism = []
        for row in rows:
            item = row["item"]
            e = ov.get(item, {})
            recon_name = e.get("name") or b.camel(item)
            recon_analyte = e.get("analyteColumn")
            recon_skip = (item in base_items) or (norm_item(item) in central) or bool(e.get("role"))
            routable = row["P"] in ROUTE_P or row["A"] in ROUTE_A

            if item in analyte_by_item:                       # ANALYTE axis (names + each $def)
                recon_names = [c["name"] for c in (recon_analyte or [])]
                if recon_names != analyte_by_item[item]:
                    mism.append(f"{item}: analyte cols real={analyte_by_item[item]} recon={recon_names}")
                else:
                    for c in (recon_analyte or []):
                        if c["name"] not in real_defs:
                            continue  # amap-only membership name, no $def emitted
                        recon_def = L.analyte_column_obj(c["name"], item, row["desc"],
                                                         c["dtype"], None, c["readOnly"])
                        if recon_def != real_defs[c["name"]]:
                            mism.append(f"{item}/{c['name']}: analyte $def differs (dtype/readOnly)")
            elif item in routed:                              # NAME axis (a routed field)
                if recon_skip:
                    mism.append(f"{item}: recon says skip but real routes it as '{name_by_item[item]}'")
                elif recon_name != name_by_item[item]:
                    mism.append(f"{item}: name real='{name_by_item[item]}' recon='{recon_name}'")
            elif routable:                                    # ANNOTATION-SKIP axis
                # content-routing would emit it, but real route() did not -> must be a
                # base_items or role-override skip in the reconstruction.
                if not recon_skip:
                    mism.append(f"{item}: real suppressed a routable row but recon has no skip")
            # else: not routable by content -> skipped both ways, nothing to check
        total += len(rows)
        fails += len(mism)
        status = "PASS" if not mism else f"FAIL ({len(mism)})"
        print(f"{tapp:16s} rows={len(rows):3d}  {status}")
        for m in mism[:12]:
            print(f"    - {m}")
    print(f"\n{'ALL PASS' if not fails else str(fails)+' MISMATCHES'} "
          f"across {total} rows in {len(TIER_TAPPS)} TAPPs.")
    print("PASS => content + sidecar reproduces route() exactly => identical generated schemas.")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
