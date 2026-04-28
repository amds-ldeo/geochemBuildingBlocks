"""Generate an xlsx dataset-entry template from a TAPP instance.

Usage:
    python tools/build_dataset_template.py <tapp-instance.json> [<out.xlsx>]

The TAPP instance's ada:analyteTemplate.ada:analyteColumns[] become the
worksheet's column headers (with the column's schema:name as the visible
header and schema:valueName as a stable key in row 2 metadata). Rows are
seeded from the instance's ada:analyteTemplate.ada:defaultAnalytes — one
row per analyte. The user fills in measured values per analyte/analysis.

If <out.xlsx> is omitted, writes alongside the input file with suffix
"-dataset-template.xlsx".
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _resolve_column_meta(col_ref: dict, base_dir: Path) -> dict:
    """A column entry in the TAPP instance can be inline OR a {@id, ...} ref.
    For now we only handle inline columns — refs would need network/file
    resolution we don't yet implement here. Returns {valueName, name, description}."""
    return {
        "valueName": col_ref.get("schema:valueName") or col_ref.get("@id", "").rsplit("/", 1)[-1],
        "name": col_ref.get("schema:name") or col_ref.get("schema:valueName") or "",
        "description": col_ref.get("schema:description") or "",
    }


def _row_value(row: dict, key: str):
    """Pull a row value by valueName (defaultAnalytes rows are
    additionalProperties keyed by analyteColumn valueName)."""
    return row.get(key)


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n", 1)[0])
    parser.add_argument("tapp_instance", help="Path to a TAPP instance JSON file.")
    parser.add_argument("out", nargs="?", default=None,
                        help="Output xlsx path. Default: <tapp_instance>-dataset-template.xlsx.")
    args = parser.parse_args()

    in_path = Path(args.tapp_instance)
    if not in_path.exists():
        sys.exit(f"input not found: {in_path}")

    with open(in_path, "r", encoding="utf-8") as f:
        tapp = json.load(f)

    template = tapp.get("ada:analyteTemplate") or {}
    cols = template.get("ada:analyteColumns") or []
    rows = template.get("ada:defaultAnalytes") or []
    if not cols:
        sys.exit(f"input has no ada:analyteTemplate.ada:analyteColumns: {in_path}")

    out_path = Path(args.out) if args.out else in_path.with_name(in_path.stem + "-dataset-template.xlsx")

    # Resolve column metadata
    col_metas = [_resolve_column_meta(c, in_path.parent) for c in cols]
    col_keys = [m["valueName"] for m in col_metas]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analyses"

    header_font = Font(bold=True)
    key_font = Font(italic=True, color="666666", size=9)
    desc_font = Font(italic=True, color="666666", size=9)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Row 1: human-readable column names
    # Row 2: schema:valueName (machine-readable key)
    # Row 3: schema:description (guidance)
    # Row 4+: data (one row per defaultAnalyte, plus blanks for new analyses)
    for i, m in enumerate(col_metas, start=1):
        ws.cell(row=1, column=i, value=m["name"]).font = header_font
        ws.cell(row=1, column=i).fill = header_fill
        ws.cell(row=2, column=i, value=m["valueName"]).font = key_font
        ws.cell(row=3, column=i, value=m["description"]).font = desc_font
        ws.cell(row=3, column=i).alignment = wrap
        # Set a reasonable column width
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, min(40, len(str(m["name"])) + 4))

    # Seed default analytes
    for ri, row in enumerate(rows, start=4):
        for ci, key in enumerate(col_keys, start=1):
            v = _row_value(row, key)
            if v is not None:
                ws.cell(row=ri, column=ci, value=v if not isinstance(v, (dict, list)) else json.dumps(v))

    # Add a few blank rows for new analyses
    blank_start = 4 + len(rows)
    for ri in range(blank_start, blank_start + 3):
        for ci in range(1, len(col_keys) + 1):
            ws.cell(row=ri, column=ci, value=None)

    # Freeze header rows (top 3)
    ws.freeze_panes = "A4"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"wrote {out_path} — {len(col_keys)} columns, {len(rows)} pre-seeded rows")


if __name__ == "__main__":
    main()
