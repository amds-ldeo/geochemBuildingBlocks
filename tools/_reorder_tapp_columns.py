"""One-shot: reorder TAPP workbook columns so structural cols (Level / CDIF path /
matchComment / implementation notes) sit in cols G-J immediately after the A-F
metadata block, with pub columns trailing in cols K..onwards.

Old layout: A-E (item..example) | F (Last update) | G..W (17 pubs) | X..AA (4 struct)
New layout: A-E (item..example) | F (Last update) | G..J (4 struct) | K..AA (17 pubs)

Preserves cell values, styles, column widths, and merged ranges (which only exist
in cols A-E and are unaffected). Run once per file:

    python tools/_reorder_tapp_columns.py docs/TAPP_EPMA_filled.xlsx
    python tools/_reorder_tapp_columns.py docs/TAPP_EPMA_filled-noInterp.xlsx
"""
from __future__ import annotations
import sys
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Old col indices (1-indexed)
PUB_OLD_START, PUB_OLD_END = 7, 23      # G..W
STRUCT_OLD_START, STRUCT_OLD_END = 24, 27  # X..AA
N_PUBS = PUB_OLD_END - PUB_OLD_START + 1   # 17
N_STRUCT = STRUCT_OLD_END - STRUCT_OLD_START + 1  # 4

# New col indices
STRUCT_NEW_START = 7   # G
PUB_NEW_START = STRUCT_NEW_START + N_STRUCT  # K = 11


def snapshot_col(ws, col_idx: int, max_row: int) -> dict:
    """Return value + style snapshot for one column."""
    cells = []
    for r in range(1, max_row + 1):
        c = ws.cell(row=r, column=col_idx)
        cells.append({
            "value": c.value,
            "font": copy(c.font) if c.has_style else None,
            "fill": copy(c.fill) if c.has_style else None,
            "border": copy(c.border) if c.has_style else None,
            "alignment": copy(c.alignment) if c.has_style else None,
            "number_format": c.number_format,
            "protection": copy(c.protection) if c.has_style else None,
            "comment": copy(c.comment) if c.comment else None,
        })
    letter = get_column_letter(col_idx)
    cd = ws.column_dimensions.get(letter)
    width = cd.width if cd else None
    hidden = cd.hidden if cd else False
    return {"cells": cells, "width": width, "hidden": hidden, "letter": letter}


def write_col(ws, col_idx: int, snap: dict) -> None:
    for r, info in enumerate(snap["cells"], start=1):
        c = ws.cell(row=r, column=col_idx)
        c.value = info["value"]
        if info["font"] is not None:
            c.font = info["font"]
        if info["fill"] is not None:
            c.fill = info["fill"]
        if info["border"] is not None:
            c.border = info["border"]
        if info["alignment"] is not None:
            c.alignment = info["alignment"]
        if info["number_format"]:
            c.number_format = info["number_format"]
        if info["protection"] is not None:
            c.protection = info["protection"]
        if info["comment"] is not None:
            c.comment = info["comment"]
    new_letter = get_column_letter(col_idx)
    if snap["width"] is not None:
        ws.column_dimensions[new_letter].width = snap["width"]
    ws.column_dimensions[new_letter].hidden = snap["hidden"]


def reorder(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb["TAPP"]
    max_row = ws.max_row

    pub_snaps = [snapshot_col(ws, c, max_row) for c in range(PUB_OLD_START, PUB_OLD_END + 1)]
    struct_snaps = [snapshot_col(ws, c, max_row) for c in range(STRUCT_OLD_START, STRUCT_OLD_END + 1)]

    # Wipe widths on all moved columns so we can reassign cleanly.
    for c in range(PUB_OLD_START, STRUCT_OLD_END + 1):
        letter = get_column_letter(c)
        if letter in ws.column_dimensions:
            del ws.column_dimensions[letter]

    # Write struct block at cols 7..10
    for i, snap in enumerate(struct_snaps):
        write_col(ws, STRUCT_NEW_START + i, snap)

    # Write pub block at cols 11..27
    for i, snap in enumerate(pub_snaps):
        write_col(ws, PUB_NEW_START + i, snap)

    wb.save(path)
    print(f"reordered {path}")


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: python tools/_reorder_tapp_columns.py <xlsx> [<xlsx>...]")
        return 2
    for arg in sys.argv[1:]:
        p = Path(arg)
        if not p.exists():
            print(f"skip (not found): {p}")
            continue
        reorder(p)
    return 0


if __name__ == "__main__":
    sys.exit(main())
