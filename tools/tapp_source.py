#!/usr/bin/env python3
"""Read a TAPP source table, whichever format it is in.

The library moved from xlsx to CSV with the 2026-08-11 delivery, and the delivery README is
explicit that the CSV is the source of truth and the xlsx a generated artifact that should not be
parsed for content. The path-driven pipeline had three separate readers, each opening the workbook
with openpyxl and re-deriving the header layout — so switching source format meant changing the
same logic three times, in three slightly different forms.

This is the one primitive they all actually wanted: the TAPP sheet as a list of row tuples, header
first. Everything above it — locating columns, skipping group headers, finding the mode block —
stays where it was.

CSV rows are normalised to look exactly like openpyxl's: padded to the header width, and empty
cells returned as None rather than ''. Callers already treat both as absent (build_tapp.norm maps
each to ""), but matching openpyxl exactly means a reader cannot behave differently on one format
than the other.

Encoding is utf-8-sig: the library's CSVs carry a BOM, and their content includes superscripts and
Greek that must survive (²⁰⁶Pb/²³⁸U, δ⁵⁶Fe, J cm⁻²).
"""
import csv
import os


def rows(path):
    """The TAPP table as a list of row tuples, header first. Accepts .csv or .xlsx."""
    if str(path).lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8-sig") as f:
            raw = [list(r) for r in csv.reader(f)]
        if not raw:
            return []
        width = len(raw[0])
        return [tuple((c if c != "" else None) for c in (r + [""] * (width - len(r)))[:width])
                for r in raw]

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = "TAPP" if "TAPP" in wb.sheetnames else wb.sheetnames[0]
    out = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return out


def exists(path):
    return bool(path) and os.path.exists(path)


_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# The canonical TAPP source library is the `tapp/` git submodule (amds-ldeo/tapp), whose root holds
# the delivery contents directly (Current TAPPs/, Claude Skills for TAPP/, composed_tapps.json) —
# no TAPPS<date>/ wrapper. The older inline TAPPS<date>/ drops are kept only as a fallback.
_TAPP_SUBMODULE = os.path.join(_ROOT, "tapp")


def _is_delivery(path):
    """True if `path` looks like a TAPP delivery (has the tables and/or the manifest)."""
    return os.path.isdir(os.path.join(path, "Current TAPPs")) or \
        os.path.exists(os.path.join(path, "composed_tapps.json"))


def current_delivery():
    """The TAPP source library.

    Prefers the `tapp/` submodule (amds-ldeo/tapp) — the canonical source going forward, version
    pinned by the submodule commit. Falls back to the newest inline TAPPS<date>/ folder for a repo
    that still carries a drop inline (deliveries are dated, so the latest name is the latest drop).

    Derived rather than hard-coded because three separate tools once pinned TAPPS20260811 in their
    own constants and silently read a superseded delivery after the next drop.
    """
    if _is_delivery(_TAPP_SUBMODULE):
        return _TAPP_SUBMODULE
    ds = sorted(d for d in os.listdir(_ROOT)
                if d.startswith("TAPPS") and os.path.isdir(os.path.join(_ROOT, d)))
    return os.path.join(_ROOT, ds[-1]) if ds else _ROOT


def manifest_path():
    """The composed_tapps.json to use, or None.

    Prefers the submodule's manifest, then the newest across any inline TAPPS<date>/ folders. Not
    simply `current_delivery()/composed_tapps.json`: some drops shipped without a manifest, so
    anything that assumed the current delivery has one crashed outright. The composition declaration
    changes far less often than the tables, so falling back to the most recent one is better than
    failing — as long as the caller says which delivery it came from.
    """
    sub = os.path.join(_TAPP_SUBMODULE, "composed_tapps.json")
    if os.path.exists(sub):
        return sub
    for d in sorted((x for x in os.listdir(_ROOT) if x.startswith("TAPPS")), reverse=True):
        p = os.path.join(_ROOT, d, "composed_tapps.json")
        if os.path.exists(p):
            return p
    return None


def modules_dir():
    """The authoritative modules folder: <newest delivery>/Claude Skills for TAPP/modules.

    Confirmed by Stephen as the source of truth for both the module CSVs and their sidecars — the
    2026-08-13 drop also ships copies under references/modules/, which are NOT authoritative.
    """
    return os.path.join(current_delivery(), "Claude Skills for TAPP", "modules")
