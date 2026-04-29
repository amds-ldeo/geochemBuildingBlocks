"""Heuristic analyte-axis inference for the publication columns of the TAPP
spreadsheet, plus a side workbook with each <pub>-interp column inserted
right after its source pub (layout A) for side-by-side review, plus paired
exampleempaTAPP-<pub>-interp.json / exampledetailEMPA-<pub>-interp.json
review files generated from the inferred data via the existing
example_for_pub builder.

For each pub column the script parses:

  Row 59 — Primary Calibration Standard Name. Standard entries shaped like
           "Anorthite (Si Kα, Al Kα, Ca Kα); Albite (Na Kα); ..." are split
           on `;` (or `,` when there's no `;`); each entry's parenthetical
           is parsed for element symbols and optional Kα/Kβ/Lα/Lβ X-ray-
           line tags so we get {element: (line, standard)}.
  Row 64 — Typical Detection Limit. Two patterns supported:
             * "<Compound>: <value>; <Compound>: <value>; …"
             * "<value> for <Compound>, <Compound>, …; <value> for …"
           Compounds are bare element symbols (Si, Mg, …) or oxide
           formulas (SiO2, Al2O3, Cr2O3, FeO, …) — element extracted from
           the oxide.
  Row 48 — Halogen Correction on Oxygen. F, Cl, OH, CO2, S mentions are
           treated as analytes when present.
  Row 32 — Target Element. If explicitly populated (pipe- or comma-delim)
           takes precedence over inference.

Outputs:
  docs/TAPP_EPMA_filled-interp.xlsx   (layout A side workbook)
  build/interp-review/exampleempaTAPP-<pub>-interp.json
  build/interp-review/exampledetailEMPA-<pub>-interp.json

The interp JSON files are rebuilt every run; review them to validate
the inferred analyte axis and per-analyte mappings before merging the
interp column data back into the source spreadsheet.
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path
from shutil import copyfile

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError):
    pass

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "docs" / "TAPP_EPMA_filled.xlsx"
OUT_XLSX = REPO / "docs" / "TAPP_EPMA_filled-interp.xlsx"
REVIEW_DIR = REPO / "build" / "interp-review"

KNOWN_ELEMENTS = set("""
H He Li Be B C N O F Ne
Na Mg Al Si P S Cl Ar
K Ca Sc Ti V Cr Mn Fe Co Ni Cu Zn Ga Ge As Se Br Kr
Rb Sr Y Zr Nb Mo Tc Ru Rh Pd Ag Cd In Sn Sb Te I Xe
Cs Ba La Ce Pr Nd Pm Sm Eu Gd Tb Dy Ho Er Tm Yb Lu Hf Ta W Re Os Ir Pt Au Hg Tl Pb Bi Po At Rn
Fr Ra Ac Th Pa U Np Pu
""".split())

VOLATILES = ("F", "Cl", "OH", "CO2", "S", "H2O")
ENTRY_RE = re.compile(r"([^();,]+?)\s*\(([^)]+)\)")
EL_LINE_RE = re.compile(
    r"([A-Z][a-z]?)"
    r"(?:\s+(K\s*alpha|K\s*beta|L\s*alpha|L\s*beta|Kα|Kβ|Lα|Lβ))?"
)
OXIDE_RE = re.compile(r"\b([A-Z][a-z]?)\d*O\d*\b")
LIMIT_FOR_RE = re.compile(r"([^;]*?)\s+for\s+([^;]+)", re.IGNORECASE)
LIMIT_KV_RE = re.compile(r"\b([A-Z][a-z]?\d*(?:[A-Z][a-z]?\d*)?)\s*[:=]\s*([^;,]+)")


# ---------- parsers ----------

def parse_calibration(text: str):
    if not text:
        return []
    out, seen = [], set()
    for m in ENTRY_RE.finditer(text):
        standard = m.group(1).strip(" ;,").strip()
        if standard.lower().startswith("for "):
            continue
        for piece in re.split(r"[,;]", m.group(2)):
            em = EL_LINE_RE.match(piece.strip())
            if not em:
                continue
            element = em.group(1)
            if element not in KNOWN_ELEMENTS:
                continue
            line = em.group(2)
            if line:
                line = (line.strip()
                        .replace("alpha", "α").replace("beta", "β")
                        .replace("K α", "Kα").replace("K β", "Kβ")
                        .replace("L α", "Lα").replace("L β", "Lβ"))
            key = (element, standard)
            if key in seen:
                continue
            seen.add(key)
            out.append((element, line, standard))
    return out


def parse_detection_limits(text: str):
    if not text:
        return []
    out, seen = [], set()
    for m in LIMIT_KV_RE.finditer(text):
        compound, value = m.group(1).strip(), m.group(2).strip()
        ox = OXIDE_RE.match(compound)
        if ox and ox.group(1) in KNOWN_ELEMENTS:
            element = ox.group(1)
        elif compound in KNOWN_ELEMENTS:
            element = compound
        else:
            continue
        if element in seen:
            continue
        seen.add(element)
        out.append((element, value))
    for m in LIMIT_FOR_RE.finditer(text):
        value = m.group(1).strip(" ;,")
        for piece in re.split(r"[,;]", m.group(2)):
            piece = piece.strip()
            ox = OXIDE_RE.match(piece)
            if ox and ox.group(1) in KNOWN_ELEMENTS:
                element = ox.group(1)
            elif piece in KNOWN_ELEMENTS:
                element = piece
            else:
                continue
            if element in seen:
                continue
            seen.add(element)
            out.append((element, value))
    return out


def parse_volatiles_row(text: str):
    if not text:
        return []
    return [v for v in VOLATILES
            if re.search(rf"(?<![A-Za-z]){re.escape(v)}(?![A-Za-z0-9])", text)]


def merge_analytes(*lists):
    out, seen = [], set()
    for lst in lists:
        for el in lst:
            if el not in seen:
                seen.add(el)
                out.append(el)
    return out


# ---------- layout A: interp column adjacent to each source pub ----------

def build_layout_a(src_path: Path, dst_path: Path, pub_cols, interp_data):
    """Write a fresh xlsx where each <pub>-interp column sits right after its
    source pub. Returns dict label → dst column index of the interp column."""
    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb["TAPP"]

    dst_wb = openpyxl.Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "TAPP"

    # Compute src→dst column mapping: copy each src col, and after each pub col,
    # leave one empty dst col for the interp.
    pub_idx_set = {c for c, _, _ in pub_cols}
    pub_idx_to_label = {c: lbl for c, lbl, _ in pub_cols}

    src_to_dst = {}
    interp_dst_col_by_label = {}
    dst_c = 1
    for src_c in range(1, src_ws.max_column + 1):
        src_to_dst[src_c] = dst_c
        dst_c += 1
        if src_c in pub_idx_set:
            interp_dst_col_by_label[pub_idx_to_label[src_c]] = dst_c
            dst_c += 1

    # Copy all source values into dst at remapped columns
    for r in range(1, src_ws.max_row + 1):
        for src_c in range(1, src_ws.max_column + 1):
            v = src_ws.cell(row=r, column=src_c).value
            if v is not None:
                dst_ws.cell(row=r, column=src_to_dst[src_c], value=v)

    # Write interp columns: header + verbatim copy of source pub + analyte-axis overrides
    for label, dst_c in interp_dst_col_by_label.items():
        dst_ws.cell(row=1, column=dst_c, value=f"{label}-interp")
        src_col = next(c for c, lbl, _ in pub_cols if lbl == label)
        for r in range(2, src_ws.max_row + 1):
            v = src_ws.cell(row=r, column=src_col).value
            if v is not None:
                dst_ws.cell(row=r, column=dst_c, value=v)
        # Override analyte-axis rows with interp data (overwrites the verbatim copy)
        for r, val in interp_data.get(label, {}).items():
            dst_ws.cell(row=r, column=dst_c, value=val)

    # Set reasonable column widths matching source for the pub columns
    for src_c, dst_c in src_to_dst.items():
        try:
            w = src_ws.column_dimensions[src_ws.cell(row=1, column=src_c).column_letter].width
            if w:
                dst_ws.column_dimensions[dst_ws.cell(row=1, column=dst_c).column_letter].width = w
        except Exception:
            pass

    dst_wb.save(dst_path)
    return interp_dst_col_by_label


# ---------- generate review JSON files from the interp columns ----------

def generate_interp_examples(layout_xlsx: Path, interp_dst_col_by_label: dict):
    """Use the existing lib's example_for_pub to generate paired (TAPP, detail)
    JSON examples from the interp columns. Files written to build/interp-review/."""
    sys.path.insert(0, str(REPO / "tools"))
    import _tapp_lib as L

    L.configure("empaTAPP", layout_xlsx)

    wb = openpyxl.load_workbook(layout_xlsx, data_only=True)
    ws = wb["TAPP"]

    # Locate structural columns by header (they shifted right after layout A)
    cdif_col = mc_col = impl_col = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if not isinstance(h, str):
            continue
        h = h.strip()
        if h == "CDIF-geochem schema path":
            cdif_col = c
        elif h == "matchComment":
            mc_col = c
        elif h == "implementation notes":
            impl_col = c

    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    written = []

    for label, src_col in interp_dst_col_by_label.items():
        # Build rows with `pubs = [interp_value]` so example_for_pub treats it as pub_index 0
        rows = []
        for r in range(2, ws.max_row + 1):
            item = ws.cell(row=r, column=1).value
            impl = ws.cell(row=r, column=impl_col).value if impl_col else None
            if item is None and impl is None:
                continue
            rec = {
                "item": str(item).strip() if item is not None else None,
                "desc": str(ws.cell(row=r, column=2).value).strip()
                        if ws.cell(row=r, column=2).value else None,
                "dtype_col": str(ws.cell(row=r, column=4).value).strip()
                             if ws.cell(row=r, column=4).value else None,
                "example": str(ws.cell(row=r, column=5).value).strip()
                           if ws.cell(row=r, column=5).value else None,
                "cdif_path": (str(ws.cell(row=r, column=cdif_col).value).strip()
                              if cdif_col and ws.cell(row=r, column=cdif_col).value else None),
                "matchComment": (str(ws.cell(row=r, column=mc_col).value).strip()
                                 if mc_col and ws.cell(row=r, column=mc_col).value else None),
                "impl": str(impl).strip() if impl is not None else None,
                "pubs": [ws.cell(row=r, column=src_col).value],
            }
            rec["parsed"] = L.parse_impl(rec["impl"])
            rows.append(rec)

        # Skip pubs that have no Target Element value (nothing to anchor on)
        target_val = next((row["pubs"][0] for row in rows if row["item"] == "Target Element"), None)
        if not target_val:
            continue

        # Override PUBS so example_for_pub uses the right code/label for @ids etc.
        L.PUBS = [(label, f"{label} (interp)")]
        empa_ex, detail_ex = L.example_for_pub(0, f"{label} (interp)", rows)

        # Adjust @id suffix to include "-interp" so the review file is clearly tagged
        empa_ex["@id"] = f"ex:empaTAPP-{label}-interp"
        detail_ex["@id"] = f"ex:detailEMPA-{label}-interp"
        detail_ex["schema:measurementTechnique"] = {"@id": empa_ex["@id"]}

        empa_path = REVIEW_DIR / f"exampleempaTAPP-{label}-interp.json"
        with open(empa_path, "w", encoding="utf-8") as f:
            json.dump(empa_ex, f, indent=2, ensure_ascii=False)
            f.write("\n")
        written.append(empa_path)
        if detail_ex.get("schema:additionalProperty"):
            detail_path = REVIEW_DIR / f"exampledetailEMPA-{label}-interp.json"
            with open(detail_path, "w", encoding="utf-8") as f:
                json.dump(detail_ex, f, indent=2, ensure_ascii=False)
                f.write("\n")
            written.append(detail_path)
    return written


# ---------- main ----------

def main():
    if not SRC.exists():
        sys.exit(f"source not found: {SRC}")

    # First pass: read source, find pub columns, compute interp data
    src_wb = openpyxl.load_workbook(SRC, data_only=True)
    src_ws = src_wb["TAPP"]

    pub_cols = []  # (col_idx, label, full_header)
    for c in range(7, src_ws.max_column + 1):
        h = src_ws.cell(row=1, column=c).value
        if h and isinstance(h, str) and re.match(r"^P\d", h.strip()):
            label = h.split("\n", 1)[0].split(":", 1)[0].strip()
            pub_cols.append((c, label, h))
    print(f"Found {len(pub_cols)} pub columns: {[p[1] for p in pub_cols]}")

    # Locate analyte-axis rows by label
    label_to_row = {}
    for r in range(2, src_ws.max_row + 1):
        item = src_ws.cell(row=r, column=1).value
        if isinstance(item, str):
            label_to_row[item.strip()] = r
    R_TARGET = label_to_row.get("Target Element")
    R_XRAY = label_to_row.get("X-ray Line")
    R_HALOGEN = label_to_row.get("Halogen Correction on Oxygen")
    R_CALIB = label_to_row.get("Primary Calibration Standard Name")
    R_DETLIM = label_to_row.get("Typical Detection Limit")

    # Build interp data per pub: {label: {row_idx: value}}
    interp_data = {}
    print()
    for col, label, _ in pub_cols:
        cal_text = src_ws.cell(row=R_CALIB, column=col).value if R_CALIB else None
        det_text = src_ws.cell(row=R_DETLIM, column=col).value if R_DETLIM else None
        hal_text = src_ws.cell(row=R_HALOGEN, column=col).value if R_HALOGEN else None
        tgt_text = src_ws.cell(row=R_TARGET, column=col).value if R_TARGET else None

        cal_entries = parse_calibration(str(cal_text) if cal_text else "")
        det_entries = parse_detection_limits(str(det_text) if det_text else "")
        vol_entries = parse_volatiles_row(str(hal_text) if hal_text else "")

        explicit = []
        if tgt_text:
            sep = "|" if "|" in str(tgt_text) else ","
            explicit = [a.strip() for a in str(tgt_text).split(sep) if a.strip()]
        analyte_list = (merge_analytes(explicit) if explicit
                        else merge_analytes([e for e, _, _ in cal_entries],
                                            [e for e, _ in det_entries],
                                            vol_entries))

        line_map = {e: ln for e, ln, _ in cal_entries}
        std_map = {e: std for e, _, std in cal_entries}
        det_map = {e: dl for e, dl in det_entries}

        per_pub = {}
        if analyte_list:
            per_pub[R_TARGET] = "|".join(analyte_list)
            if R_XRAY:
                per_pub[R_XRAY] = "|".join(line_map.get(e) or "" for e in analyte_list)
            if R_CALIB:
                per_pub[R_CALIB] = "|".join(std_map.get(e) or "" for e in analyte_list)
            if R_DETLIM:
                per_pub[R_DETLIM] = "|".join(det_map.get(e) or "" for e in analyte_list)
        interp_data[label] = per_pub
        print(f"  {label}: analytes = {analyte_list}")

    # Layout A: side xlsx with interp columns adjacent
    interp_cols = build_layout_a(SRC, OUT_XLSX, pub_cols, interp_data)
    print(f"\nwrote {OUT_XLSX} (layout A: interp columns adjacent to source pubs)")

    # Generate review JSON files from interp columns
    written = generate_interp_examples(OUT_XLSX, interp_cols)
    print(f"\nwrote {len(written)} review JSON files under {REVIEW_DIR.relative_to(REPO)}/")
    for p in written:
        print(f"  {p.relative_to(REPO)}")


if __name__ == "__main__":
    main()
