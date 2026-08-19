#!/usr/bin/env python3
"""Find in-band 'not applicable' sentinels in the TAPP workbooks' allowed-content column.

The application Y/N grid (see validate_application_grid.py) will generate absence constraints: an
N means the generated schema REQUIRES the property to be absent. But the workbooks predate that
mechanism, so authors expressed the same idea inside the allowed content — 'N/A (not
multi-collector)'. Once the grid is enforced the two can contradict each other, so the in-band
sentinels need triaging first. That is what this does.

Not every N/A is the same thing, so classify before counting:

  BARE       a delimiter-flanked 'N/A' / 'None' in an enumeration tail
             ('... | Other: specify | Unknown | N/A | None') — the house convention on every
             controlled list, says nothing about applications. Noise; hidden unless --all.

  QUALIFIED  an N/A carrying an explanation ('N/A (not multi-collector)') — the author saying WHY
             it does not apply. These interact with the grid, and split again by what they blame:

    APPLICATION  the reason names an application/mode of this technique. If the row is already N
                 for those applications the sentinel is REDUNDANT and should be dropped; if the
                 row is all-Y the grid may be missing an N, or the distinction may have no column
                 (LA-Q_SF-ICPMS gates on sampling geometry, so its analyser-dependent rows have
                 nowhere to go and legitimately stay in-band).

    INSTRUMENT   the reason names a hardware capability ('instrument does not have guard
                 electrode'). This varies BETWEEN instruments within one application, so the grid
                 can never express it — always stays in-band.

Beware the mixed case, which the classifier reports as APPLICATION but a human must split:

    'Not applicable (SF-ICP-MS, or MC-ICP-MS without a collision/reaction cell, or STD mode)'

The first clause duplicates the grid; the rest are within-application and must survive.

Usage:
    python tools/scan_na_sentinels.py                            # every workbook
    python tools/scan_na_sentinels.py docs/TEM_TAPP_v7.xlsx      # one
    python tools/scan_na_sentinels.py --all                      # include the BARE tail matches
"""
import argparse
import glob
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# the application block ends at whichever of these appears first (matches validate_application_grid)
END_HEADERS = ("schema path", "literature assessment")
SECTION_RE = re.compile(r"^\d+\.\s")

# an N/A-ish token plus whatever qualifies it, up to the next pipe delimiter
NA_RE = re.compile(r"(?:\bN\s*/\s*A\b|\bnot applicable\b|\bnone\b)([^|]*)", re.I)

# hardware capability, not a mode of working
INSTRUMENT_RE = re.compile(
    r"instrument (?:does not|doesn't|lacks)|not (?:installed|fitted|present|equipped)"
    r"|only\b|hardware|geometric magnification|no \w+ (?:installed|fitted)", re.I)


def classify(content):
    """[(kind, snippet)] for each N/A-ish token in one allowed-content cell."""
    hits = []
    for m in NA_RE.finditer(content):
        qual = m.group(1).strip(" \t'\"")
        # bare when nothing qualifies it, or the qualifier is just the next list item
        if len(qual) < 3 or not qual.startswith(("(", "—", "-", ":")):
            hits.append(("BARE", m.group(0).strip()))
            continue
        kind = "INSTRUMENT" if INSTRUMENT_RE.search(qual) else "APPLICATION"
        hits.append((kind, re.sub(r"\s+", " ", m.group(0).strip())))
    return hits


def scan(xlsx):
    """[(item, [Y apps], [N apps], [(kind, snippet)])] or None when there is no grid to compare."""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "TAPP" not in wb.sheetnames:
        wb.close()
        return None
    rows = list(wb["TAPP"].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None
    hdr = [(str(c).strip() if c is not None else "") for c in rows[0]]

    start = next((i + 1 for i, h in enumerate(hdr) if h.lower().startswith("last update")), None)
    end = next((i for i, h in enumerate(hdr) if h.lower() in END_HEADERS), None)
    ex = next((i for i, h in enumerate(hdr) if h.lower().startswith("example")), None)
    if start is None or end is None or end <= start or ex is None:
        return None
    apps = hdr[start:end]

    out = []
    for r in rows[1:]:
        item = (str(r[0]).strip() if r[0] else "")
        if not item or SECTION_RE.match(item):
            continue                 # blank line or section header, not a property
        content = re.sub(r"\s+", " ", str(r[ex]).strip()) if ex < len(r) and r[ex] else ""
        hits = classify(content)
        if not hits:
            continue
        vals = [(str(r[i]).strip().upper() if i < len(r) and r[i] is not None else "")
                for i in range(start, end)]
        present = [v for v in vals if v in ("Y", "N")]
        if not present:
            continue                 # not a Y/N grid (see validate_application_grid)
        on = [a for a, v in zip(apps, vals) if v == "Y"]
        off = [a for a, v in zip(apps, vals) if v == "N"]
        out.append((item, on, off, hits))
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", nargs="?", help="one workbook (default: all TAPP workbooks)")
    ap.add_argument("--all", action="store_true",
                    help="also list the BARE enumeration-tail matches (noisy)")
    args = ap.parse_args()

    if args.workbook:
        targets = [args.workbook if os.path.isabs(args.workbook)
                   else os.path.join(ROOT, args.workbook)]
    else:
        targets = sorted(set(glob.glob(os.path.join(ROOT, "docs", "*_TAPP_*.xlsx"))) |
                         set(glob.glob(os.path.join(ROOT, "docs", "TAPP_*.xlsx"))))
        targets = [t for t in targets
                   if not os.path.basename(t).startswith("~$")      # Excel lock file
                   and "template" not in os.path.basename(t).lower()
                   and "interp" not in os.path.basename(t).lower()]

    keep = ("BARE", "APPLICATION", "INSTRUMENT") if args.all else ("APPLICATION", "INSTRUMENT")
    counts = {"BARE": 0, "APPLICATION": 0, "INSTRUMENT": 0}
    for t in targets:
        name = os.path.basename(t)
        if not os.path.exists(t):
            print(f"missing: {t}")
            continue
        res = scan(t)
        if res is None:
            print(f"{name:<44s} no grid / no allowed-content column — skipped")
            continue
        shown = [(i, on, off, [h for h in hits if h[0] in keep]) for i, on, off, hits in res]
        shown = [s for s in shown if s[3]]
        for _, _, _, hits in res:
            for kind, _snippet in hits:
                counts[kind] += 1
        if not shown:
            print(f"{name:<44s} clean")
            continue
        print(f"{name}")
        for item, on, off, hits in shown:
            for kind, snippet in hits:
                gate = "conditional" if off else "ALL-Y"
                print(f"   {kind:<11s} [{gate}] {item}")
                print(f"        Y: {', '.join(on) or '(none)'}")
                if off:
                    print(f"        N: {', '.join(off)}")
                print(f"        {snippet[:110]}")

    print()
    print(f"APPLICATION-blamed: {counts['APPLICATION']}   "
          f"INSTRUMENT-blamed: {counts['INSTRUMENT']}   "
          f"bare enumeration tails: {counts['BARE']}"
          + ("" if args.all else " (hidden — pass --all)"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
